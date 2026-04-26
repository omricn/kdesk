import json
import logging

from django.contrib import messages
from django.shortcuts import redirect, render

logger = logging.getLogger(__name__)


# ── Excel → HTML parser ───────────────────────────────────────────────────────

# Office default theme color palette (indices 0–9)
_THEME_COLORS = [
    'FFFFFF', '000000', 'E7E6E6', '44546A',
    '4472C4', 'ED7D31', 'A5A5A5', 'FFC000',
    '5B9BD5', '70AD47',
]

_BORDER_STYLES = {
    'thin': '1px solid', 'medium': '2px solid', 'thick': '3px solid',
    'hair': '1px solid', 'dashed': '1px dashed', 'dotted': '1px dotted',
    'double': '3px double', 'mediumDashed': '2px dashed',
    'dashDot': '1px dashed', 'slantDashDot': '1px dashed',
}


def _tint(hex6, tint):
    r, g, b = int(hex6[0:2], 16), int(hex6[2:4], 16), int(hex6[4:6], 16)
    if tint >= 0:
        r, g, b = (round(x + (255 - x) * tint) for x in (r, g, b))
    else:
        r, g, b = (round(x * (1 + tint)) for x in (r, g, b))
    return '{:02X}{:02X}{:02X}'.format(
        min(255, max(0, r)), min(255, max(0, g)), min(255, max(0, b))
    )


def _color_css(c):
    if c is None:
        return None
    try:
        t = getattr(c, 'tint', 0) or 0
        if c.type == 'rgb':
            argb = c.rgb or '00000000'
            if len(argb) < 8 or argb[:2] == '00':
                return None
            hex6 = argb[2:8].upper()
            return f'#{_tint(hex6, t) if t else hex6}'
        if c.type == 'theme':
            idx = getattr(c, 'theme', 0) or 0
            if 0 <= idx < len(_THEME_COLORS):
                hex6 = _THEME_COLORS[idx]
                return f'#{_tint(hex6, t) if t else hex6}'
    except Exception:
        pass
    return None


def _border_css(side):
    if not side or not getattr(side, 'border_style', None):
        return None
    style = _BORDER_STYLES.get(side.border_style, '1px solid')
    color = _color_css(getattr(side, 'color', None)) or '#adb5bd'
    return f'{style} {color}'


def _fmt_value(cell):
    from django.utils.html import escape
    from datetime import datetime, date as date_type

    v = cell.value
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'TRUE' if v else 'FALSE'
    if isinstance(v, (int, float)):
        nf = (cell.number_format or '').strip()
        if '%' in nf:
            return f'{v * 100:.1f}%'
        if any(s in nf for s in ('$', '€', '£', '₪', r'\$')):
            return f'{v:,.2f}'
        if '#,##' in nf or '_(' in nf:
            return (f'{int(v):,}' if v == int(v) else f'{v:,.2f}')
        return f'{v:g}'
    if isinstance(v, (datetime, date_type)):
        return v.strftime('%d/%m/%Y')
    return escape(str(v))


def excel_to_sheets_html(file_obj):
    """Parse workbook; return list of {name, html} for every visible sheet."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    result = []

    for ws in wb.worksheets:
        # Merged cell map: top-left → (rowspan, colspan); rest → skip
        spans, skip = {}, set()
        for mr in ws.merged_cells.ranges:
            r1, c1, r2, c2 = mr.min_row, mr.min_col, mr.max_row, mr.max_col
            spans[(r1, c1)] = (r2 - r1 + 1, c2 - c1 + 1)
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    if (r, c) != (r1, c1):
                        skip.add((r, c))

        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        buf = ['<table class="budget-table"><colgroup>']

        for ci in range(1, max_col + 1):
            letter = get_column_letter(ci)
            dim = ws.column_dimensions.get(letter)
            if dim and dim.hidden:
                buf.append('<col class="col-hidden">')
            elif dim and dim.width:
                buf.append(f'<col style="width:{max(30, int(dim.width * 7.5))}px;">')
            else:
                buf.append('<col>')

        buf.append('</colgroup><tbody>')

        for ri in range(1, max_row + 1):
            rdim = ws.row_dimensions.get(ri)
            if rdim and rdim.hidden:
                continue
            row_h = (f' style="height:{max(16, int(rdim.height * 1.33))}px;"'
                     if rdim and rdim.height else '')
            buf.append(f'<tr{row_h}>')

            for ci in range(1, max_col + 1):
                letter = get_column_letter(ci)
                cdim = ws.column_dimensions.get(letter)
                if (cdim and cdim.hidden) or (ri, ci) in skip:
                    continue

                cell = ws.cell(row=ri, column=ci)
                styles = []

                try:
                    fill = cell.fill
                    if fill and fill.fill_type not in (None, 'none'):
                        bg = _color_css(fill.fgColor)
                        if bg and bg.upper() not in ('#FFFFFF', '#FFFFFF'):
                            styles.append(f'background-color:{bg};')
                except Exception:
                    pass

                try:
                    font = cell.font
                    if font:
                        if font.bold:
                            styles.append('font-weight:bold;')
                        if font.italic:
                            styles.append('font-style:italic;')
                        if font.underline and font.underline != 'none':
                            styles.append('text-decoration:underline;')
                        fc = _color_css(font.color)
                        if fc and fc.upper() != '#000000':
                            styles.append(f'color:{fc};')
                        if font.size and font.size != 11:
                            styles.append(f'font-size:{font.size}pt;')
                except Exception:
                    pass

                try:
                    al = cell.alignment
                    if al:
                        if al.horizontal in ('center', 'right', 'left'):
                            styles.append(f'text-align:{al.horizontal};')
                        if al.vertical == 'center':
                            styles.append('vertical-align:middle;')
                        elif al.vertical == 'top':
                            styles.append('vertical-align:top;')
                        if al.wrap_text:
                            styles.append('white-space:pre-wrap;word-break:break-word;')
                except Exception:
                    pass

                try:
                    border = cell.border
                    if border:
                        for side_name in ('top', 'right', 'bottom', 'left'):
                            css = _border_css(getattr(border, side_name))
                            if css:
                                styles.append(f'border-{side_name}:{css};')
                except Exception:
                    pass

                style_attr = f' style="{" ".join(styles)}"' if styles else ''
                span_attr = ''
                if (ri, ci) in spans:
                    rs, cs = spans[(ri, ci)]
                    if rs > 1:
                        span_attr += f' rowspan="{rs}"'
                    if cs > 1:
                        span_attr += f' colspan="{cs}"'

                buf.append(f'<td{style_attr}{span_attr}>{_fmt_value(cell)}</td>')

            buf.append('</tr>')

        buf.append('</tbody></table>')
        result.append({'name': ws.title, 'html': ''.join(buf)})

    wb.close()
    return result


# ── View ──────────────────────────────────────────────────────────────────────

def budget_view(request):
    from tickets.views import admin_required  # reuse existing decorator logic
    if not request.user.is_authenticated or not request.user.is_superuser:
        from django.contrib import messages as msg
        msg.error(request, 'Access denied.')
        return redirect('dashboard')

    from .models import BudgetFile

    if request.method == 'POST':
        uploaded = request.FILES.get('excel_file')
        if not uploaded:
            messages.error(request, 'No file selected.')
            return redirect('budget')
        if not uploaded.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, 'Only .xlsx and .xls files are supported.')
            return redirect('budget')

        # Delete old files so only one budget file exists at a time
        BudgetFile.objects.all().delete()

        bf = BudgetFile(
            original_name=uploaded.name,
            uploaded_by=request.user,
            is_processing=True,
        )
        bf.file.save(uploaded.name, uploaded)
        bf.save()

        from .tasks import parse_budget_file
        parse_budget_file.delay(bf.pk)

        messages.success(request, f'"{uploaded.name}" uploaded — processing in background, page will refresh automatically.')
        return redirect('budget')

    budget_file = BudgetFile.objects.first()
    sheets = []
    if budget_file and budget_file.rendered_sheets:
        try:
            sheets = json.loads(budget_file.rendered_sheets)
        except Exception:
            pass

    return render(request, 'budget/budget.html', {
        'budget_file': budget_file,
        'sheets': sheets,
    })
